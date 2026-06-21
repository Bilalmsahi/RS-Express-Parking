from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from .models import Booking, BookingUser
from django.utils import timezone
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from datetime import datetime
from django.utils.translation import gettext_lazy as _
import io
import zipfile
from decimal import Decimal
from django.http import HttpResponse
from payments.views import generate_invoice, generate_booking_form_pdf
from django.contrib.admin import FieldListFilter
from django.db.models import Q, Func, Value, CharField
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import timedelta
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill
import pytz
from django.db.models.functions import Lower
from payments.views import send_booking_confirmation_email, send_manager_new_booking_email
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

class SupplierFilter(FieldListFilter):
    """Custom filter for suppliers showing 'All Suppliers' and individual supplier options."""
    title = 'Supplier'

    def __init__(self, field, request, params, model, model_admin, field_path):
        self.lookup_val = request.GET.get(field_path)
        super().__init__(field, request, params, model, model_admin, field_path)

    def expected_parameters(self):
        return [self.field_path]

    def choices(self, changelist):
        from users.models import Supplier
        
        # "All" choice - no filter
        yield {
            'selected': self.lookup_val is None,
            'query_string': changelist.get_query_string(remove=[self.field_path]),
            'display': 'All'
        }

        # "All Suppliers" choice - only bookings with a supplier
        yield {
            'selected': self.lookup_val == 'all_suppliers',
            'query_string': changelist.get_query_string({self.field_path: 'all_suppliers'}, []),
            'display': 'All Suppliers (Any)'
        }

        # Individual supplier choices
        for supplier in Supplier.objects.filter(is_active=True).order_by('name'):
            yield {
                'selected': self.lookup_val == str(supplier.id),
                'query_string': changelist.get_query_string({self.field_path: supplier.id}, []),
                'display': f'{supplier.name} ({supplier.booking_id_prefix})'
            }

    def queryset(self, request, queryset):
        if self.lookup_val == 'all_suppliers':
            return queryset.filter(supplier__isnull=False)
        elif self.lookup_val:
            return queryset.filter(supplier_id=self.lookup_val)
        return queryset


class ServiceFilter(FieldListFilter):
    """Custom filter for services. Suppliers only see RS/Both services, not Dublin-only."""
    title = 'Service'

    def __init__(self, field, request, params, model, model_admin, field_path):
        self.lookup_val = request.GET.get(field_path)
        self.request = request
        super().__init__(field, request, params, model, model_admin, field_path)

    def expected_parameters(self):
        return [self.field_path]

    def choices(self, changelist):
        from services.models import Service
        
        # "All" choice - no filter
        yield {
            'selected': self.lookup_val is None,
            'query_string': changelist.get_query_string(remove=[self.field_path]),
            'display': 'All'
        }

        # Get services based on user role
        if getattr(self.request.user, 'is_supplier', False):
            # Suppliers only see RS Express or Both services
            services = Service.objects.exclude(website__in=['dublinairportparking', 'none']).order_by('name')
        else:
            # Other users see all services
            services = Service.objects.order_by('name')

        for service in services:
            yield {
                'selected': self.lookup_val == str(service.id),
                'query_string': changelist.get_query_string({self.field_path: service.id}, []),
                'display': service.name
            }

    def queryset(self, request, queryset):
        if self.lookup_val:
            return queryset.filter(service_id=self.lookup_val)
        return queryset


class WebsiteFilter(FieldListFilter):
    """Custom filter for website that excludes supplier bookings when filtered."""
    title = 'Website'

    def __init__(self, field, request, params, model, model_admin, field_path):
        self.lookup_val = request.GET.get(field_path)
        self.request = request
        self.model = model
        super().__init__(field, request, params, model, model_admin, field_path)

    def expected_parameters(self):
        return [self.field_path]

    def choices(self, changelist):
        # "All" choice - no filter
        yield {
            'selected': self.lookup_val is None,
            'query_string': changelist.get_query_string(remove=[self.field_path]),
            'display': 'All'
        }

        websites = (
            self.model.objects
            .exclude(website__isnull=True)
            .exclude(website__exact='')
            .values_list('website', flat=True)
            .distinct()
            .order_by('website')
        )

        for website in websites:
            yield {
                'selected': self.lookup_val == website,
                'query_string': changelist.get_query_string({self.field_path: website}, []),
                'display': website
            }

    def queryset(self, request, queryset):
        if self.lookup_val:
            return queryset.filter(website=self.lookup_val, supplier__isnull=True)
        return queryset


class SingleDateFilter(FieldListFilter):
    template = 'admin/single_date_filter.html'

    def __init__(self, field, request, params, model, model_admin, field_path):
        self.field_path = field_path
        self.date_params = {
            'gte': f'{field_path}__gte',
            'lt': f'{field_path}__lt'
        }
        if 'departure_time' in field_path:
            field.verbose_name = 'Departure Date'
        elif 'return_time' in field_path:
            field.verbose_name = 'Return Date'
        elif 'created_at' in field_path:
            field.verbose_name = 'Created Date'
        super().__init__(field, request, params, model, model_admin, field_path)

    def expected_parameters(self):
        return [self.date_params['gte'], self.date_params['lt']]

    def queryset(self, request, queryset):
        filters = []
        for field in ['departure_time', 'return_time', 'created_at']:
            gte = request.GET.get(f"{field}__gte")
            lt = request.GET.get(f"{field}__lt")
            q = Q()
            if gte:
                try:
                    start_datetime = timezone.make_aware(datetime.strptime(gte, '%Y-%m-%d'))
                    q &= Q(**{f"{field}__gte": start_datetime})
                except Exception:
                    pass
            if lt:
                try:
                    end_datetime = timezone.make_aware(datetime.strptime(lt, '%Y-%m-%d')) + timedelta(days=1)
                    q &= Q(**{f"{field}__lt": end_datetime})
                except Exception:
                    pass
            if q:
                filters.append(q)
        if filters:
            query = filters[0]
            for f in filters[1:]:
                query |= f
            return queryset.filter(query).distinct()
        return queryset

    def value(self):
        value = self.used_parameters.get(self.date_params['gte'])
        if isinstance(value, list):
            return value[0] if value else ''
        return value or ''

    def choices(self, changelist):
        yield {
            'selected': self.value() is None,
            'query_string': changelist.get_query_string(remove=self.expected_parameters()),
            'display': _('All')
        }
        if self.value():
            yield {
                'selected': True,
                'query_string': changelist.get_query_string(),
                'display': f'Selected Date: {self.value()}'
            }

class RemoveSpaces(Func):
    function = 'REPLACE'
    arity = 3

    def __init__(self, expression):
        super().__init__(expression, Value(' '), Value(''), output_field=CharField())  

class RemoveChars(Func):
    function = 'REPLACE'
    arity = 3

    def __init__(self, expression, char_to_remove):
        super().__init__(expression, Value(char_to_remove), Value(''), output_field=CharField())          

IRELAND_TZ = pytz.timezone('Europe/Dublin')
@admin.register(Booking)
class BookingAdmin(admin.ModelAdmin):

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Check for active date filters and order accordingly
        departure_start = request.GET.get('departure_time__gte')
        departure_end = request.GET.get('departure_time__lt')
        return_start = request.GET.get('return_time__gte')
        return_end = request.GET.get('return_time__lt')
        created_start = request.GET.get('created_at__gte')
        created_end = request.GET.get('created_at__lt')

        # Suppliers only see their own bookings
        if getattr(request.user, 'is_supplier', False):
            qs = qs.filter(supplier=request.user.supplier_profile)

        # Hide "pending" and "payment failed" for managers
        if request.user.is_manager:
            qs = qs.exclude(status__name__iexact="pending").exclude(status__name__iexact="payment failed")

        if departure_start or departure_end:
            return qs.order_by('departure_time')
        elif return_start or return_end:
            return qs.order_by('return_time')
        elif created_start or created_end:
            return qs.order_by('created_at')
        return qs

    def price_display(self, obj):
        if obj.discounted_price is not None and obj.discounted_price < obj.total_price:
            return format_html(
                '<span style="text-decoration: line-through; color: #888;">€{}</span><br>'
                '<span style="color: #28a745; font-weight: bold;">€{}</span>',
                obj.total_price, obj.discounted_price
            )
        return f"€{obj.total_price}"

    price_display.short_description = 'Total Price'

    def supplier_price_display(self, obj):
        extension = Decimal(str(obj.supplier_extended_amount or 0))
        total_price = max(Decimal(str(obj.total_price)) - extension, Decimal('0.00'))
        discounted_price = None
        if obj.discounted_price is not None:
            discounted_price = max(Decimal(str(obj.discounted_price)) - extension, Decimal('0.00'))

        if discounted_price is not None and discounted_price < total_price:
            return format_html(
                '<span style="text-decoration: line-through; color: #888;">€{}</span><br>'
                '<span style="color: #28a745; font-weight: bold;">€{}</span>',
                total_price, discounted_price
            )
        return f"€{total_price}"

    supplier_price_display.short_description = 'Total Price'


    def get_search_results(self, request, queryset, search_term):
        # Normalize terms
        search_term_no_spaces = search_term.replace(' ', '')
        search_term_no_dash = search_term.replace('-', '')

        # Run default admin search on the already-filtered queryset (includes date filters)
        base_qs, use_distinct = super().get_search_results(request, queryset, search_term)

        # Build pk subqueries for custom searches (avoid queryset unions)
        ors = [Q(pk__in=base_qs.values_list('pk', flat=True))]

        if search_term_no_spaces:
            reg_ids = self.model.objects.annotate(
                reg_no_no_spaces=RemoveSpaces('car_registration_no')
            ).filter(
                reg_no_no_spaces__icontains=search_term_no_spaces
            ).values_list('pk', flat=True)
            ors.append(Q(pk__in=reg_ids))

        if search_term_no_dash:
            bid_ids = self.model.objects.annotate(
                booking_id_normalized=Lower(RemoveChars('booking_id', '-'))
            ).filter(
                booking_id_normalized__icontains=search_term_no_dash.lower()
            ).values_list('pk', flat=True)
            ors.append(Q(pk__in=bid_ids))

        combined_q = Q()
        for part in ors:
            combined_q |= part

        # Apply on the base filtered queryset (preserves date filters and others)
        final_qs = queryset.filter(combined_q).distinct()
        return final_qs, True

    def download_invoices_zip(self, request, queryset):
        zip_buffer = io.BytesIO()
        today = timezone.now().strftime('%d-%m-%Y')
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for booking in queryset:
                response, error = generate_invoice(booking)
                if error:
                    continue
                pdf_bytes = response.content
                filename = f"Invoice-{booking.booking_id or booking.id}.pdf"
                zip_file.writestr(filename, pdf_bytes)
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="invoices_{today}.zip"'
        return response

    download_invoices_zip.short_description = "Download invoices for selected bookings (ZIP)"

    def download_booking_forms_zip(self, request, queryset):
        zip_buffer = io.BytesIO()
        today = timezone.now().strftime('%d-%m-%Y')
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for booking in queryset:
                response, error = generate_booking_form_pdf(booking)
                if error:
                    continue
                pdf_bytes = response.content
                filename = f"BookingForm-{booking.booking_id or booking.id}.pdf"
                zip_file.writestr(filename, pdf_bytes)
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="booking_forms_{today}.zip"'
        return response

    download_booking_forms_zip.short_description = "Download booking forms for selected bookings (ZIP)"

    def colored_status(self, obj):
        status_colors = {
            'completed': '#28a745',
            'confirmed': '#ffc107',
            'cancelled': '#dc3545',
            'payment failed': '#dc3545',
            'rescheduled': '#fd7e14',
            'pending': 'gray',
            'started': "#3552b1",
        }
        status_html = ''

        if obj.status:
            status_name = obj.status.name.lower()
            color = status_colors.get(status_name, 'inherit')
            status_html += format_html(
                '<div><span style="background-color: {}; color: #fff; padding: 4px 8px; '
                'border-radius: 4px; font-weight: bold; display: inline-block;">{}</span></div>',
                color, obj.status.name
            )
        
        ireland_today = timezone.now().astimezone(IRELAND_TZ).date()
        departure_time_ireland = timezone.localtime(obj.departure_time, IRELAND_TZ).date()
        created_at_ireland = timezone.localtime(obj.created_at, IRELAND_TZ).date()
        if departure_time_ireland == ireland_today and created_at_ireland == ireland_today:
            status_html += format_html(
                '<div style="margin-top: 6px;">'
                '<span style="background: #e53935; color: #fff; padding: 4px 8px; border-radius: 4px; '
                'font-size: 1.1em; font-weight: bold; text-transform: uppercase; letter-spacing: 1px; '
                'display: inline-block;">TODAY</span></div>'
            )

        if not obj.is_viewed:
            status_html += format_html(
                '<div style="margin-top: 6px;"><span style="background: linear-gradient(45deg, #ff6b6b, #ee0979); '
                'color: #fff; padding: 2px 8px; border-radius: 4px; '
                'box-shadow: 0 2px 4px rgba(238,9,121,0.15); font-size: 1em; '
                'font-weight: bold; text-transform: uppercase; letter-spacing: 0.5px; display: inline-block;">NEW</span></div>'
            )
            
        return format_html(status_html) if status_html else '-'
    
    colored_status.short_description = 'Status'

    def change_view(self, request, object_id, form_url='', extra_context=None):
        # Mark the booking as viewed when opening its details
        obj = self.get_object(request, object_id)
        if obj and not obj.is_viewed:
            obj.is_viewed = True
            obj.save()
        return super().change_view(request, object_id, form_url, extra_context)

    def number(self, obj):  # Change method name from 'id' to 'number'
        return obj.id
    number.short_description = 'No'
    number.admin_order_field = 'id'

    def departure_time_display(self, obj):
        return timezone.localtime(obj.departure_time).strftime('%d %b %Y %H:%M')
    departure_time_display.admin_order_field = 'departure_time'
    departure_time_display.short_description = 'Departure Time'

    def return_time_display(self, obj):
        return timezone.localtime(obj.return_time).strftime('%d %b %Y %H:%M')
    return_time_display.admin_order_field = 'return_time'
    return_time_display.short_description = 'Return Time'

    def created_at_display(self, obj):
        return timezone.localtime(obj.created_at).strftime('%d %b %Y %H:%M')
    created_at_display.admin_order_field = 'created_at'
    created_at_display.short_description = 'Created At'

    def created_at_field(self, obj):
        return timezone.localtime(obj.created_at).strftime('%d %b %Y %H:%M')
    created_at_field.short_description = 'Created At'

    def modified_at_field(self, obj):
        return timezone.localtime(obj.modified_at).strftime('%d %b %Y %H:%M')
    modified_at_field.short_description = 'Modified At'

    def terminal_display(self, obj):
        return format_html(
            '{}<br><span style="color: #888;">{}</span>',
            obj.departure_terminal, obj.return_terminal
        )
    terminal_display.short_description = 'Terminal'

    def customer_display(self, obj):
        return format_html(
            '{} {}<br><span style="color: #888;">{}</span>',
            obj.first_name, obj.last_name, obj.contact_no
        )
    customer_display.short_description = 'Customer'

    def vehicle_display(self, obj):
        return format_html(
            '{} {}<br><span style="color: #888;">{}, {}</span>',
            obj.car_manufacturer, obj.car_model, obj.car_colour, obj.car_registration_no
        )
    vehicle_display.short_description = 'Car Registration'

    def service_display(self, obj):
        service_name = obj.service.name if obj.service else 'N/A'
        addons = ', '.join([addon.name for addon in obj.add_ons.all()])
        if addons:
            return format_html('{}<br><span style="color: #888;">{}</span>', service_name, addons)
        return service_name
    service_display.short_description = 'Service & Add-ons'

    def website_abbr(self, obj):
        if getattr(obj, 'supplier_id', None) is not None:
            return obj.supplier.booking_id_prefix.upper() if obj.supplier and obj.supplier.booking_id_prefix else "-"
        if obj.website == "rsexpressparking":
            return "RS"
        elif obj.website == "dublinairportparking":
            return "DAEP"
        return obj.website or "-"
    website_abbr.short_description = "Website"

    def get_list_display(self, request):
        if getattr(request.user, 'is_supplier', False):
            return ('number', 'booking_id', 'customer_display', 'departure_time_display', 'return_time_display', 'service_display','vehicle_display', 'terminal_display', 
                'supplier_price_display', 'colored_status', 'created_at_display', 'download_invoice', 'download_booking_form')
        if request.user.is_manager:
            return ('number', 'booking_id', 'customer_display', 'departure_time_display', 'return_time_display', 'service_display', 'vehicle_display', 'terminal_display', 'colored_status', 'created_at_display', 'download_booking_form')
        return ('number', 'booking_id', 'customer_display', 'departure_time_display', 'return_time_display', 'service_display','vehicle_display', 'terminal_display', 
                'price_display', 'colored_status', 'created_at_display', 'download_invoice', 'download_booking_form', 'website_abbr')
    

    def get_fieldsets(self, request, obj=None):
        base_fieldsets = [
            ('Customer Information', {
                'fields': (
                    ('first_name', 'last_name'),
                    ('email', 'contact_no'),
                    ('total_passengers'),
                )
            }),
            ('Vehicle Details', {
                'fields': (
                    ('car_registration_no', 'car_model'),
                    ('car_manufacturer', 'car_colour')
                )
            }),
            ('Travel Information', {
                'fields': (
                    ('departure_terminal', 'return_terminal'),
                    ('departure_flight_number', 'return_flight_number'),
                    ('departure_time', 'return_time')
                )
            }),
            ('Service & Add-ons', {
                'fields': (
                    'service',
                    'add_ons',
                )
            }),
        ]

        if not request.user.is_manager:
            is_super_admin = (
                request.user.is_superuser
                and not getattr(request.user, 'is_manager', False)
                and not getattr(request.user, 'is_supplier', False)
            )
            pricing_fields = ('total_price', 'coupon', 'discounted_price', 'bonus_points_used', 'previous_price')
            if obj and obj.supplier and is_super_admin:
                pricing_fields = pricing_fields + ('supplier_extended_amount',)
            base_fieldsets.append(('Pricing', {
                'fields': (
                    pricing_fields,
                )
            }))

        # Add status section with appropriate fields based on user type
        supplier_profile = getattr(request.user, 'supplier_profile', None) if getattr(request.user, 'is_supplier', False) else None
        supplier_uses_custom_ids = bool(supplier_profile and supplier_profile.use_custom_booking_ids)

        status_fields = ['status', 'note']
        if obj or supplier_uses_custom_ids:
            status_fields.insert(0, 'booking_id')
        if obj:
            # For existing bookings, show the formatted display
            status_fields.extend([('created_at_field', 'modified_at_field')])
        else:
            # For new bookings, show the editable field
            status_fields.extend(['created_at'])

        if not request.user.is_manager:
            status_fields.append('download_invoice')
        status_fields.append('download_booking_form')
        
        base_fieldsets.append(('Booking Status', {
            'fields': tuple(status_fields)
        }))

        return base_fieldsets
    
    def get_list_filter(self, request):
        base_filters = [
            ('departure_time', SingleDateFilter),
            ('return_time', SingleDateFilter),
            ('created_at', SingleDateFilter),
            'status',
            ('service', ServiceFilter),
            ('website', WebsiteFilter),
            ('supplier', SupplierFilter),
        ]
        # Remove 'website' filter for managers and suppliers
        if getattr(request.user, "is_manager", False) or getattr(request.user, "is_supplier", False):
            return [
                ('departure_time', SingleDateFilter),
                ('return_time', SingleDateFilter),
                ('created_at', SingleDateFilter),
                'status',
                ('service', ServiceFilter),
            ]
        return base_filters
    search_fields = ('first_name', 'last_name', 'email', 'car_registration_no', 'contact_no', 'booking_id', 'id')

    def get_readonly_fields(self, request, obj=None):
        base = ('modified_at_field', 'download_invoice', 'download_booking_form')
        if obj:  # If editing existing object
            base = base + ('created_at_field', 'booking_id')
        if getattr(request.user, 'is_supplier', False):
            base = base + ('supplier',)
        return base
    
    list_per_page = 20 
    show_full_result_count = True  # Shows total number of items
    

    def full_name(self, obj):
        return f"{obj.first_name} {obj.last_name}"
    full_name.short_description = 'Customer Name'

    def download_invoice(self, obj):
        if obj.booking_id:
            url = reverse('invoice-download', args=[obj.booking_id])
            return format_html(
                '<a class="button" href="{}" target="_blank">'
                '<img src="/static/admin/img/icon-yes.svg" alt="Print Invoice"> Print Invoice</a>',
                url
            )
        return '-'

    download_invoice.short_description = 'Invoice'
    download_invoice.allow_tags = True

    def download_booking_form(self, obj):
        if obj.booking_id:
            url = reverse('booking-form-download', args=[obj.booking_id])
            return format_html(
                '<a class="button" href="{}" target="_blank">'
                '<img src="/static/admin/img/icon-yes.svg" alt="Print Form"> Booking Form</a>',
                url
            )
        return '-'

    download_booking_form.short_description = 'Download Booking Form'
    download_booking_form.allow_tags = True


    actions = ['export_as_csv', 'export_as_pdf', 'download_invoices_zip', 'download_booking_forms_zip']

    def get_actions(self, request):
        actions = super().get_actions(request)
        if request.user.is_manager:
            # Remove the download_invoices_zip action for managers
            actions.pop('download_invoices_zip', None)
        if getattr(request.user, 'is_supplier', False):
            # Remove delete action for suppliers
            actions.pop('delete_selected', None)
        return actions

    def has_delete_permission(self, request, obj=None):
        """
        Prevent suppliers from deleting bookings.
        """
        if getattr(request.user, 'is_supplier', False):
            return False
        return super().has_delete_permission(request, obj)
    
    def get_dynamic_filename(self, departure_start, departure_end, return_start, return_end, created_start, created_end):
        def format_range(label, start, end):
            if start and end and start != end:
                start_str = datetime.strptime(start, '%Y-%m-%d').strftime('%d_%b')
                end_str = datetime.strptime(end, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_{start_str}-{end_str}"
            elif start and end and start == end:
                start_str = datetime.strptime(start, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_{start_str}"
            elif start:
                start_str = datetime.strptime(start, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_from_{start_str}"
            elif end:
                end_str = datetime.strptime(end, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_until_{end_str}"
            return ""

        parts = []
        dep = format_range("Departures", departure_start, departure_end)
        arr = format_range("Arrivals", return_start, return_end)
        cre = format_range("Created", created_start, created_end)
        for part in (dep, arr, cre):
            if part:
                parts.append(part)

        # Use year from the first available date, fallback to current year
        for date in (departure_start, departure_end, return_start, return_end, created_start, created_end):
            if date:
                year = datetime.strptime(date, '%Y-%m-%d').strftime('%Y')
                break
        else:
            year = datetime.now().strftime('%Y')

        if parts:
            return f"{'_&_'.join(parts)}_{year}.xlsx"
        else:
            return "Selected_Bookings.xlsx"
        
    def filter_by_range(self, qs, field, start, end):
        if start:
            qs = qs.filter(**{f"{field}__gte": timezone.make_aware(datetime.strptime(start, '%Y-%m-%d'))})
        if end:
            qs = qs.filter(**{f"{field}__lt": timezone.make_aware(datetime.strptime(end, '%Y-%m-%d')) + timedelta(days=1)})
        return qs

    def export_as_csv(self, request, queryset):
        """
        Export bookings organized by active date filters (now supports start/end for each)
        """
        wb = Workbook()
        ws = wb.active

        # Get active filters (start and end for each)
        departure_start = request.GET.get('departure_time__gte')
        departure_end = request.GET.get('departure_time__lt')
        return_start = request.GET.get('return_time__gte')
        return_end = request.GET.get('return_time__lt')
        created_start = request.GET.get('created_at__gte')
        created_end = request.GET.get('created_at__lt')

        headers = [
            'PRODUCT CODE','Booking Ref', 'Customer/Contact', 'Entry Date & Time', 'Return Date & Time', 'Vehicle/Registration',
            'Out Flight Number', 'In Flight Number', 'Add Onns'
        ]

        current_row = 1

        def write_section(title, filtered_queryset, start_row):
            # Write section title
            title_cell = ws.cell(row=start_row, column=1, value=title)
            title_cell.font = Font(bold=True, size=24, color="FFFFFF")
            title_cell.alignment = Alignment(vertical='center', horizontal='center')
            title_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=start_row, column=col)
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws.row_dimensions[start_row].height = 40

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row + 1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Write data
            for idx, booking in enumerate(filtered_queryset, start=start_row + 2):
                row_data = [
                    booking.service.name if booking.service else '',
                    booking.booking_id,                 
                    f'{booking.first_name} {booking.last_name} {booking.contact_no}',                
                    timezone.localtime(booking.departure_time).strftime('%d %b %Y %H:%M'),
                    timezone.localtime(booking.return_time).strftime('%d %b %Y %H:%M'),
                    f'{booking.car_manufacturer} {booking.car_model} {booking.car_colour} {booking.car_registration_no}',
                    booking.departure_flight_number,
                    booking.return_flight_number,
                    ', '.join([addon.name for addon in booking.add_ons.all()]),
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=idx, column=col, value=value)
                    ws.cell(row=idx, column=col).alignment = Alignment(horizontal='left')
            return start_row + filtered_queryset.count() + 4

        def format_date_display(start, end):
            if start and end:
                return f"{datetime.strptime(start, '%Y-%m-%d').strftime('%d %b')} - {datetime.strptime(end, '%Y-%m-%d').strftime('%d %b')}"
            elif start:
                return f"from {datetime.strptime(start, '%Y-%m-%d').strftime('%d %b')}"
            elif end:
                return f"until {datetime.strptime(end, '%Y-%m-%d').strftime('%d %b')}"
            return ""

        # Write sections for each filter if either start or end is set
        if departure_start or departure_end:
            departure_qs = self.filter_by_range(queryset, 'departure_time', departure_start, departure_end).order_by('departure_time')
            if departure_qs.exists():
                current_row = write_section(f"Departures {format_date_display(departure_start, departure_end)}", departure_qs, current_row)
                current_row += 1

        if return_start or return_end:
            return_qs = self.filter_by_range(queryset, 'return_time', return_start, return_end).order_by('return_time')
            if return_qs.exists():
                current_row = write_section(f"Returns {format_date_display(return_start, return_end)}", return_qs, current_row)
                current_row += 1

        if created_start or created_end:
            created_qs = self.filter_by_range(queryset, 'created_at', created_start, created_end).order_by('created_at')
            if created_qs.exists():
                current_row = write_section(f"Created {format_date_display(created_start, created_end)}", created_qs, current_row)

        if not (departure_start or departure_end or return_start or return_end or created_start or created_end):
            if queryset.exists():
                current_row = write_section("All Selected", queryset.order_by('created_at'), current_row)

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                cell = row[col_idx - 1]
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            if max_length > 0:
                ws.column_dimensions[column_letter].width = max_length + 2

        filename = self.get_dynamic_filename(departure_start, departure_end, return_start, return_end, created_start, created_end)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response    

    export_as_csv.short_description = "Export Selected Bookings as Excel"

    def get_dynamic_pdf_filename(self, departure_start, departure_end, return_start, return_end, created_start, created_end):
        """Generate dynamic filename for PDF export (same logic as Excel but with .pdf extension)"""
        def format_range(label, start, end):
            if start and end and start != end:
                start_str = datetime.strptime(start, '%Y-%m-%d').strftime('%d_%b')
                end_str = datetime.strptime(end, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_{start_str}-{end_str}"
            elif start and end and start == end:
                start_str = datetime.strptime(start, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_{start_str}"
            elif start:
                start_str = datetime.strptime(start, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_from_{start_str}"
            elif end:
                end_str = datetime.strptime(end, '%Y-%m-%d').strftime('%d_%b')
                return f"{label}_until_{end_str}"
            return ""

        parts = []
        dep = format_range("Departures", departure_start, departure_end)
        arr = format_range("Arrivals", return_start, return_end)
        cre = format_range("Created", created_start, created_end)
        for part in (dep, arr, cre):
            if part:
                parts.append(part)

        for date in (departure_start, departure_end, return_start, return_end, created_start, created_end):
            if date:
                year = datetime.strptime(date, '%Y-%m-%d').strftime('%Y')
                break
        else:
            year = datetime.now().strftime('%Y')

        if parts:
            return f"{'_&_'.join(parts)}_{year}.pdf"
        else:
            return "Selected_Bookings.pdf"

    def export_as_pdf(self, request, queryset):
        """
        Export bookings as PDF organized by active date filters
        """
        # Get active filters (start and end for each)
        departure_start = request.GET.get('departure_time__gte')
        departure_end = request.GET.get('departure_time__lt')
        return_start = request.GET.get('return_time__gte')
        return_end = request.GET.get('return_time__lt')
        created_start = request.GET.get('created_at__gte')
        created_end = request.GET.get('created_at__lt')

        def format_date_display(start, end):
            if start and end:
                return f"{datetime.strptime(start, '%Y-%m-%d').strftime('%d %b')} - {datetime.strptime(end, '%Y-%m-%d').strftime('%d %b')}"
            elif start:
                return f"from {datetime.strptime(start, '%Y-%m-%d').strftime('%d %b')}"
            elif end:
                return f"until {datetime.strptime(end, '%Y-%m-%d').strftime('%d %b')}"
            return ""

        def prepare_booking_row(booking, cell_style):
            return [
                Paragraph(booking.service.name if booking.service else '', cell_style),
                Paragraph(booking.booking_id or '', cell_style),
                Paragraph(f'{booking.first_name} {booking.last_name} {booking.contact_no}', cell_style),
                Paragraph(timezone.localtime(booking.departure_time).strftime('%d %b %Y %H:%M'), cell_style),
                Paragraph(timezone.localtime(booking.return_time).strftime('%d %b %Y %H:%M'), cell_style),
                Paragraph(f'{booking.car_manufacturer} {booking.car_model} {booking.car_colour} {booking.car_registration_no}', cell_style),
                Paragraph(booking.departure_flight_number or '', cell_style),
                Paragraph(booking.return_flight_number or '', cell_style),
                Paragraph(', '.join([addon.name for addon in booking.add_ons.all()]), cell_style),
            ]

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=10*mm,
            rightMargin=10*mm,
            topMargin=10*mm,
            bottomMargin=10*mm
        )

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.white,
            alignment=TA_CENTER,
            spaceAfter=0,
            spaceBefore=0,
        )
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=7,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=7,
            fontName='Helvetica',
            alignment=TA_LEFT,
            leading=8,
        )

        elements = []
        headers = ['PRODUCT CODE', 'Booking Ref', 'Customer/Contact', 'Entry Date & Time', 
                   'Return Date & Time', 'Vehicle/Registration', 'Out Flight Number', 
                   'In Flight Number', 'Add Onns']
        header_row = [Paragraph(h, header_style) for h in headers]

        # Column widths (total ~277mm for A4 landscape with margins)
        col_widths = [38*mm, 28*mm, 42*mm, 30*mm, 30*mm, 50*mm, 20*mm, 20*mm, 19*mm]

        # Build sections
        if departure_start or departure_end:
            departure_qs = self.filter_by_range(queryset, 'departure_time', departure_start, departure_end).order_by('departure_time')
            if departure_qs.exists():
                # Section title
                title_table = Table([[Paragraph(f"Departures {format_date_display(departure_start, departure_end)}", title_style)]], colWidths=[sum(col_widths)])
                title_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.black), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6)]))
                elements.append(title_table)
                
                # Data table
                data = [header_row] + [prepare_booking_row(b, cell_style) for b in departure_qs]
                table = Table(data, colWidths=col_widths, repeatRows=1)
                table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]))
                # Alternate row colors
                for i in range(1, len(data)):
                    if i % 2 == 0:
                        table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), colors.Color(0.97, 0.97, 0.97))]))
                elements.append(table)
                elements.append(Spacer(1, 12))

        if return_start or return_end:
            return_qs = self.filter_by_range(queryset, 'return_time', return_start, return_end).order_by('return_time')
            if return_qs.exists():
                title_table = Table([[Paragraph(f"Returns {format_date_display(return_start, return_end)}", title_style)]], colWidths=[sum(col_widths)])
                title_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.black), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6)]))
                elements.append(title_table)
                
                data = [header_row] + [prepare_booking_row(b, cell_style) for b in return_qs]
                table = Table(data, colWidths=col_widths, repeatRows=1)
                table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]))
                for i in range(1, len(data)):
                    if i % 2 == 0:
                        table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), colors.Color(0.97, 0.97, 0.97))]))
                elements.append(table)
                elements.append(Spacer(1, 12))

        if created_start or created_end:
            created_qs = self.filter_by_range(queryset, 'created_at', created_start, created_end).order_by('created_at')
            if created_qs.exists():
                title_table = Table([[Paragraph(f"Created {format_date_display(created_start, created_end)}", title_style)]], colWidths=[sum(col_widths)])
                title_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.black), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6)]))
                elements.append(title_table)
                
                data = [header_row] + [prepare_booking_row(b, cell_style) for b in created_qs]
                table = Table(data, colWidths=col_widths, repeatRows=1)
                table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]))
                for i in range(1, len(data)):
                    if i % 2 == 0:
                        table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), colors.Color(0.97, 0.97, 0.97))]))
                elements.append(table)
                elements.append(Spacer(1, 12))

        if not (departure_start or departure_end or return_start or return_end or created_start or created_end):
            if queryset.exists():
                title_table = Table([[Paragraph("All Selected", title_style)]], colWidths=[sum(col_widths)])
                title_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.black), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6)]))
                elements.append(title_table)
                
                data = [header_row] + [prepare_booking_row(b, cell_style) for b in queryset.order_by('created_at')]
                table = Table(data, colWidths=col_widths, repeatRows=1)
                table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]))
                for i in range(1, len(data)):
                    if i % 2 == 0:
                        table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), colors.Color(0.97, 0.97, 0.97))]))
                elements.append(table)

        # Build PDF
        doc.build(elements)
        
        # Generate filename
        filename = self.get_dynamic_pdf_filename(departure_start, departure_end, return_start, return_end, created_start, created_end)

        # Create response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    export_as_pdf.short_description = "Export Selected Bookings as PDF"

    def save_model(self, request, obj, form, change):
        # Auto-set supplier for supplier users
        if getattr(request.user, 'is_supplier', False) and hasattr(request.user, 'supplier_profile'):
            obj.supplier = request.user.supplier_profile
        is_new = not change
        super().save_model(request, obj, form, change)
        if is_new and obj.status and obj.status.name.lower() == "confirmed" and not obj.email_sent and obj.supplier is None:
            try:
                send_booking_confirmation_email(obj)
                # Generate invoice PDF for manager email
                response, error = generate_invoice(obj)
                if not error:
                    pdf_content = response.getvalue()
                    send_manager_new_booking_email(obj, pdf_content)
            except Exception as e:
                self.message_user(request, f"Error sending confirmation/manager email: {e}", level='error')


    class Media:
        css = {
            'all': ('admin/css/custom_admin.css','admin/css/clickable_rows.css',)
        }
        js = ('admin/js/print_invoice.js','admin/js/clickable_rows_bookings.js')

@admin.register(BookingUser)
class BookingUserAdmin(admin.ModelAdmin):
    def created_at_display(self, obj):
        return timezone.localtime(obj.created_at).strftime('%d %b %Y %H:%M')
    created_at_display.admin_order_field = 'created_at'
    created_at_display.short_description = 'Created At'

    list_display = ('id', 'first_name', 'last_name', 'email', 'booking_completion_status', 'created_at_display')
    search_fields = ('first_name', 'last_name', 'email', 'contact_no', 'booking_id')
    list_filter = ('booking_completion_status', 'created_at',)
    readonly_fields = ('created_at_display',)

    def get_fieldsets(self, request, obj=None):
        fieldsets = [
            (None, {
                'fields': ('first_name', 'last_name', 'email', 'contact_no', 'booking_completion_status')
            }),
            ('Created At', {
                'fields': ('created_at_display',)
            }),
        ]
        return fieldsets
    
    def has_module_permission(self, request):
        """
        Restrict access for managers and suppliers.
        """
        user = request.user
        if user.is_authenticated and getattr(user, "is_manager", False):
            return False
        if user.is_authenticated and getattr(user, "is_supplier", False):
            return False
        return super().has_module_permission(request)

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css','admin/css/clickable_rows.css',)
        }
        js = ('admin/js/print_invoice.js','admin/js/clickable_rows.js')

